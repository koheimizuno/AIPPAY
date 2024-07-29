import io
from openpyxl import Workbook, load_workbook
import mojimoji

import web_util
import common_util

def get_template():
    """
    テンプレートのExcelファイル（のバイナリー）を取得
    """
    lang = web_util.get_ui_texts()

    # 新規ワークブックを生成
    wb = Workbook()
    ws = wb.active
    ws.title = lang['Pages']['Import']['TEXT000001']

    columns = 5

    # ヘッダーの生成
    ws.cell(row=1, column=1).value = lang['Pages']['Import']['TEXT000002']
    ws.cell(row=1, column=2).value = lang['Pages']['Import']['TEXT000003']
    ws.cell(row=1, column=3).value = lang['Pages']['Import']['TEXT000009']
    ws.cell(row=1, column=4).value = lang['Pages']['Import']['TEXT000010']
    ws.cell(row=1, column=5).value = lang['Pages']['Import']['TEXT000011']

    # 入力例の設定
    for r in range(4):
        ws.cell(row=2+r, column=1).value = lang['Law'][('Patent', 'Utility', 'Design', 'Trademark')[r]]
        ws.cell(row=2+r, column=2).value = '1234567'
        ws.cell(row=2+r, column=1+columns).value = lang['Pages']['Import']['TEXT000004']

    # ワークブックをバイナリー形式にして返す
    with io.BytesIO() as buff:
        wb.save(buff)
        return buff.getvalue()

def workbook_to_list(file, country='JP'):
    """
    Excelワークブックを展開して登録リストにする
    """
    lang = web_util.get_ui_texts()

    # ワークブックを読み込む
    with io.BytesIO(file) as fin:
        wb = load_workbook(fin)
        ws = wb.active

    # 項目と列番号のマッピング
    indexes = {}

    # 1行目をヘッダーとみなして、マッピングを判定
    for col in ws.iter_rows(min_row=1, max_row=1):

        for i in range(len(col)):

            n = col[i].value

            if n == lang['Pages']['Import']['TEXT000002']:
                indexes['Law'] = i
                continue

            if n == lang['Pages']['Import']['TEXT000003']:
                indexes['RegistrationNumber'] = i
                continue

            if n == lang['Pages']['Import']['TEXT000009']:
                indexes['MailAddress'] = i
                continue

            if n == lang['Pages']['Import']['TEXT000010']:
                indexes['UserOrganization'] = i
                continue

            if n == lang['Pages']['Import']['TEXT000011']:
                indexes['UserName'] = i
                continue

    if not 'Law' in indexes or not 'RegistrationNumber' in indexes:
        return None, lang['Pages']['Import']['TEXT000005']

    if not 'MailAddress' in indexes:
        return None, lang['Pages']['Import']['TEXT000005']

    # 結果のコンテナー
    result = []

    # 2行目以降をデータとして読み込む
    for row in ws.iter_rows(min_row=2):

        # 明細
        data = {}

        # 項目ごとに処理
        for key in indexes:

            i = indexes[key]

            if len(row) < i:
                continue

            # セルの値を取得
            v = row[i].value

            if v is None:
                continue

            # 法域の値変換
            if key == 'Law':
                v = str(v)
                if v in ("特許", "特許権",):
                    v = "Patent"
                if v in ("実用新案", "実用新案権", "実用新案登録", "登録実用新案",):
                    v = "Utility"
                if v in ("意匠", "意匠権", "意匠登録", "登録意匠",):
                    v = "Design"
                if v in ("商標", "商標権", "商標登録", "登録商標",):
                    v = "Trademark"
                v = v.lower()
                v = v[0].upper() + v[1:]
            
            # 登録番号の正規化
            if key == 'RegistrationNumber':
                v = str(v)
                v = mojimoji.zen_to_han(v)
                if country == 'JP':
                    if len(v) < 7 and v != "":
                        v = ('0' * int(7 - len(v))) + v

            # 値を保持
            data[key] = v

        # 結果に追加（空行は無視）
        if len(data) > 0:
            result.append(data)

    # 0件チェック
    if len(result) == 0:
        return None, lang['Pages']['Import']['TEXT000008']

    # 取得した結果を返す
    return result, None

if __name__ == '__main__':
    #data = get_template()
    #with open('out2.xlsx', 'wb') as fout:
    #    fout.write(data)
    with open('out2.xlsx', 'rb') as fin:
        data = fin.read()
        print(workbook_to_list(data))
